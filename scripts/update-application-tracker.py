#!/usr/bin/env python3
"""Maintain application-tracker.xlsx (repo root): one sheet tab per month, Status dropdown.

Used by the applying-to-job skill when it registers an application in tracker.md,
and by standalone status updates ("I got a screen", "they rejected me").

Commands:
  add         Insert one application row on the sheet for its month (creates the
              workbook / month sheet as needed). Re-running with the same company
              + role in the same month updates that row instead of duplicating.
  set-status  Change the Status dropdown value of an existing row.

Examples:
  python update-application-tracker.py add --date 2026-07-22 --company "Specsavers Canada" \
      --role "Data Analyst" --jd-file path/to/jd.md
  python update-application-tracker.py set-status --company "FGF Brands" --status Interview
"""

import argparse
import datetime as dt
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "application-tracker.xlsx"
HEADERS = ["Date", "Company Name", "Job Role", "Status", "Salary", "Job Description"]
COLUMN_WIDTHS = [12, 30, 42, 14, 30, 90]
STATUS_OPTIONS = ["Applied", "Interview", "Rejected"]
FONT_NAME = "Arial"
MAX_JD_CHARS = 32000  # Excel cell hard limit is 32,767 characters
ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
README_SHEET = "README"


def month_sheet_name(date: dt.date) -> str:
    return f"{date.strftime('%B')} {date.year}"


def parse_month_sheet_name(title: str):
    try:
        return dt.datetime.strptime(title, "%B %Y").date()
    except ValueError:
        return None


def strip_markdown(text: str) -> str:
    """Convert jd.md-style markdown to plain readable text for the Excel cell."""
    text = re.sub(r"(?m)^\s*```.*$", "", text)                      # code fences
    text = re.sub(r"!\[([^\]]*)\]\([^)]*\)", r"\1", text)           # images -> alt text
    text = re.sub(r"\[([^\]]+)\]\([^)]*\)", r"\1", text)            # [text](url) -> text
    text = re.sub(r"\*\*([^*\n]+)\*\*|__([^_\n]+)__",
                  lambda m: m.group(1) or m.group(2), text)          # bold
    text = re.sub(r"\*([^*\n]+)\*", r"\1", text)                    # *italic*
    text = re.sub(r"(?<!\w)_([^_\n]+)_(?!\w)", r"\1", text)         # _italic_
    text = re.sub(r"(?m)^\s{0,3}#{1,6}\s+", "", text)               # heading markers
    text = re.sub(r"(?m)^\s{0,3}>\s?", "", text)                    # blockquote markers
    text = re.sub(r"(?m)^\s{0,3}([-*_])\s*(?:\1\s*){2,}$", "", text)  # horizontal rules
    text = re.sub(r"(?m)^(\s*)[-*+]\s+", r"\1- ", text)             # normalize bullets
    text = text.replace("`", "")                                    # inline code ticks
    text = re.sub(r"\\([\\*_{}\[\]()#+.!>-])", r"\1", text)         # markdown escapes
    text = re.sub(r"(?im)^\s*job description \(verbatim\)\s*\n", "", text, count=1)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def clean_text(text: str) -> str:
    text = ILLEGAL_XLSX_CHARS.sub("", text or "").strip()
    if len(text) > MAX_JD_CHARS:
        text = text[:MAX_JD_CHARS] + "\n[truncated]"
    return text


def make_readme(wb: Workbook):
    ws = wb.active if wb.sheetnames == ["Sheet"] else wb.create_sheet(README_SHEET, 0)
    ws.title = README_SHEET
    lines = [
        ("Application Tracker", True),
        ("", False),
        ("One sheet tab per month; each row is one job application.", False),
        ("Rows are inserted automatically by the applying-to-job skill "
         "(scripts/update-application-tracker.py) when an application is registered.", False),
        ("Status is the only column meant to be edited by hand - use its dropdown: "
         "Applied / Interview / Rejected.", False),
        ("Salary is the posted range from the JD when the posting names one, otherwise "
         "a researched estimate marked (est., <source>).", False),
        ("The Job Description column holds the posting as plain text; the original "
         "markdown and all application artifacts live on git branch <slug> under "
         "applications/<slug>/.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(name=FONT_NAME, size=14 if bold else 11, bold=bold)
    ws.column_dimensions["A"].width = 110


def ensure_month_sheet(wb: Workbook, date: dt.date):
    title = month_sheet_name(date)
    if title in wb.sheetnames:
        return wb[title]

    # Keep month tabs in chronological order (README stays first).
    index = len(wb.sheetnames)
    months = [(parse_month_sheet_name(n), i) for i, n in enumerate(wb.sheetnames)]
    for month, i in months:
        if month and month > date.replace(day=1):
            index = i
            break
    ws = wb.create_sheet(title, index)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS[col - 1]
    ws.freeze_panes = "A2"

    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUS_OPTIONS) + '"',
        allow_blank=True,
        errorTitle="Invalid status",
        error="Pick one of: " + ", ".join(STATUS_OPTIONS),
    )
    ws.add_data_validation(dv)
    dv.add(f"D2:D1000")
    return ws


def load_or_create_workbook() -> Workbook:
    if WORKBOOK_PATH.exists():
        return load_workbook(WORKBOOK_PATH)
    wb = Workbook()
    make_readme(wb)
    return wb


def next_data_row(ws) -> int:
    row = ws.max_row
    while row > 1 and all(ws.cell(row=row, column=c).value in (None, "") for c in range(1, 7)):
        row -= 1
    return row + 1


def write_row(ws, row: int, date: dt.date, company: str, role: str, status: str,
              salary: str, jd: str):
    values = [date, company, role, status, salary, jd]
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(name=FONT_NAME, size=11)
        cell.alignment = Alignment(vertical="top", wrap_text=False)
    ws.cell(row=row, column=1).number_format = "yyyy-mm-dd"


def cmd_add(args):
    date = dt.date.fromisoformat(args.date)
    jd = ""
    if args.jd_file:
        jd = Path(args.jd_file).read_text(encoding="utf-8", errors="replace")
    elif args.jd:
        jd = args.jd
    jd = clean_text(strip_markdown(jd))
    company, role = args.company.strip(), args.role.strip()

    wb = load_or_create_workbook()
    ws = ensure_month_sheet(wb, date)

    # Same company + role in the same month: update in place, never duplicate.
    target = None
    for row in range(2, ws.max_row + 1):
        if (str(ws.cell(row=row, column=2).value or "").casefold() == company.casefold()
                and str(ws.cell(row=row, column=3).value or "").casefold() == role.casefold()):
            target = row
            break
    action = "Updated existing" if target else "Added"
    target = target or next_data_row(ws)

    write_row(ws, target, date, company, role, args.status, (args.salary or "").strip(), jd)
    wb.save(WORKBOOK_PATH)
    print(f"{action} row {target} on sheet '{ws.title}': {company} - {role} [{args.status}]")


def cmd_set_status(args):
    if not WORKBOOK_PATH.exists():
        sys.exit(f"ERROR: {WORKBOOK_PATH} does not exist - nothing to update.")
    wb = load_workbook(WORKBOOK_PATH)
    company = args.company.strip().casefold()
    role = (args.role or "").strip().casefold()

    matches = []  # newest sheet first, then latest row first
    month_sheets = sorted(
        (s for s in wb.sheetnames if parse_month_sheet_name(s)),
        key=parse_month_sheet_name, reverse=True,
    )
    for name in month_sheets:
        ws = wb[name]
        for row in range(ws.max_row, 1, -1):
            row_company = str(ws.cell(row=row, column=2).value or "").casefold()
            row_role = str(ws.cell(row=row, column=3).value or "").casefold()
            if company in row_company and (not role or role in row_role):
                matches.append((ws, row))

    if not matches:
        sys.exit(f"ERROR: no row found matching company '{args.company}'"
                 + (f" and role '{args.role}'" if args.role else ""))
    distinct = {(str(ws.cell(row=r, column=2).value), str(ws.cell(row=r, column=3).value))
                for ws, r in matches}
    if len(distinct) > 1:
        listing = "\n".join(f"  - {c} - {r}" for c, r in sorted(distinct))
        sys.exit(f"ERROR: '{args.company}' matches {len(distinct)} applications - "
                 f"narrow it with --role:\n{listing}")

    ws, row = matches[0]
    cell = ws.cell(row=row, column=4)
    cell.value = args.status
    cell.font = Font(name=FONT_NAME, size=11)
    wb.save(WORKBOOK_PATH)
    print(f"Set status '{args.status}' for {ws.cell(row=row, column=2).value} - "
          f"{ws.cell(row=row, column=3).value} (sheet '{ws.title}', row {row})")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)

    p_add = sub.add_parser("add", help="insert (or update) one application row")
    p_add.add_argument("--date", required=True, help="application date, YYYY-MM-DD")
    p_add.add_argument("--company", required=True)
    p_add.add_argument("--role", required=True)
    p_add.add_argument("--status", default="Applied", choices=STATUS_OPTIONS)
    p_add.add_argument("--salary", default="",
                       help="posted range from the JD, or researched estimate marked "
                            "(est., <source>), or 'Not posted'")
    p_add.add_argument("--jd-file", help="path to the saved jd.md (preferred)")
    p_add.add_argument("--jd", help="job description text, if no file")
    p_add.set_defaults(func=cmd_add)

    p_set = sub.add_parser("set-status", help="update the Status of an existing row")
    p_set.add_argument("--company", required=True, help="company name (substring match)")
    p_set.add_argument("--role", help="job role, to disambiguate (substring match)")
    p_set.add_argument("--status", required=True, choices=STATUS_OPTIONS)
    p_set.set_defaults(func=cmd_set_status)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
